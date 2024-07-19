import os
import time
import DrissionPage
import requests
import pandas as pd
import re
import openpyxl
import tkinter as tk
import tkinter.messagebox
import cv2
import numpy as np
import pyautogui
from threading import Thread
from PIL import Image, ImageTk


""" 飞书考勤 """
# pyinstaller -F -w FeishuCrawlaer.py


# 下载表格 =============================================================================================================
def download_excel(excel_download_path, excel_output_path, start_date, end_date):
    TRY_COUNT = 3  # 最大重试次数
    edge_path = r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe'  # 请改为你电脑内Chrome可执行文件路径
    co = DrissionPage.ChromiumOptions().set_browser_path(edge_path)  # 设置浏览器配置, 浏览器路径

    page = DrissionPage.ChromiumPage(co)  # 打开浏览器

    # 新增 tab 跳转到 url
    url = r"https://oa.feishu.cn/attendance/manage/statistics/report"
    main_tab = page.new_tab()
    main_tab.get(url)  # 跳转到 url 链接

    # 判断登录, 点击考勤
    for _ in range(TRY_COUNT):
        try:
            # 点击考勤管理按钮
            ele_考勤管理_按钮 = main_tab.ele('xpath://*[@class="ud__loading__container"]/div[4]/div[2]/div/div[1]', timeout=2)
            ele_考勤管理_按钮.click()
            break
        except Exception as e:
            # 提示登录
            try:
                main_tab.run_js('alert(arguments[0]+arguments[1]+arguments[2]);', '正在运行考勤机器人\n', ' 请在此页面直接登录飞书账号!\n',
                                '30秒后自动重试')
            except Exception as e:
                time.sleep(5)
            time.sleep(5)

    # 配置 cookie, json
    cookie = ""
    for i in main_tab.cookies(as_dict=False, all_domains=True):
        if i["name"] in ["lobsession_306", "sl_session"]:
            print(i)
            cookie += f"{i['name']}={i['value']};"
    json = {
        "query_filter": {
            "start_date": start_date,
            "end_date": end_date,
            "report_id": "102",
            "use_cache": False,
            "use_backend_date": False,
            "only_save_option": False
        }
    }

    # request 请求 task_key
    for _ in range(TRY_COUNT):
        try:
            response = requests.post(url="https://oa.feishu.cn/attendance/v2/admin/datacenter/custom_report/download",
                                     json=json, headers={'Cookie': cookie}).json()
            task_key = response["data"]["task_key"]
            break
        except Exception as e:
            pass

    # 拼接 url, 配置 headers
    url = rf"https://oa.feishu.cn/attendance/v2/admin/datacenter/download_excel?file_key={task_key}"

    # 请求文件
    for _ in range(TRY_COUNT):
        try:
            time.sleep(3)
            response = requests.get(url=url, headers={'Cookie': cookie})
            if len(response.content) > 999:
                break
        except Exception as e:
            pass

    # 下载文件
    with open(excel_download_path, 'wb') as f:
        for chunk in response.iter_content(chunk_size=1024):  # 每次请求数据不超过 chunk_size 的值, 防治大文件下载失败
            if chunk:
                f.write(chunk)
    time.sleep(2)


# 数据操作 ==============================================================================================================
def change_excel(excel_download_path, excel_output_path, start_date, end_date):
    # 读取下载的 excel 表格
    data_records = pd.read_excel(excel_download_path, skiprows=1).to_dict(orient="records")

    # 处理数据为新数据格式
    new_data_list = []
    for row_data in data_records:
        new_row_data = {"姓名": row_data["Unnamed: 0"]}
        pattern = r'^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01])$'  # 正则表达式，匹配一个字母
        for key, value in row_data.items():
            if bool(re.match(pattern, key[:10])):
                new_row_data[key] = value
        new_data_list.append(new_row_data)

    # 统计数据
    for row_data in new_data_list:
        for key, value in row_data.items():
            if key == "姓名":  # 跳过姓名列
                continue
            # 拆分出字段
            part_1 = value.split(";")[0]
            morning = part_1.split(",")[0]
            evening = part_1.split(",")[1] if len(part_1.split(",")) == 2 else ""
            part_2 = value.split(";")[1] if len(value.split(";")) == 2 else ""

            if "休息" in value:
                row_data[key] = "休"
                continue
            elif "-" == value:
                row_data[key] = ""
                continue
            elif "病假" in part_2:
                row_data[key] = "病假"
                continue
            elif "事假" in part_2:
                row_data[key] = "事假"
                continue
            elif "年假" in part_2:
                row_data[key] = "年假"
                continue
            elif "调休假" in part_2:
                row_data[key] = "调休假"
                continue

            elif "缺卡" in part_1 or "迟到" in part_1 or "早退" in part_1:
                row_data[key] = ""
                if "迟到" in morning:
                    row_data[key] += "迟到,"
                if "缺卡" in morning:
                    row_data[key] += "上班缺卡,"
                if "缺卡" in evening:
                    row_data[key] += "下班缺卡,"
                if "早退" in evening:
                    row_data[key] += "早退,"
                continue
            elif "出差" in part_2:
                row_data[key] = "出差"
                continue
            elif "正常" or "外勤" or "入职日" in morning and "正常" or "外勤" or "入职日" in evening:
                row_data[key] = "打卡"
                continue

    pd.DataFrame.from_records(new_data_list).to_excel(excel_output_path, index=False)

    def dec_to_alphanumeric(number):
        """
        数字转字母:
        1 - A; 2 - B; 4 - C; 4 - D
        :param number: 需要被转换的数字
        :return: 转换后的字母
        # dec_to_alphanumeric(1)  return A
        """
        base26 = []
        while number > 0:
            number, remainder = divmod(number, 26)
            if remainder == 0:
                base26.append('Z')
                number -= 1
            else:
                base26.append(chr(remainder - 1 + 65))
        base26 = base26[::-1]
        return ''.join(base26)

    # 加载已有的Excel文件, 进行颜色填充和批注添加
    workbook = openpyxl.load_workbook(excel_output_path)
    sheet = workbook.active  # 选择工作表

    row_index = 0  # 行索引
    for row in sheet.iter_rows(values_only=True):  # 遍历每一行
        row_index += 1  # 行索引自增, 从1开始
        cell_index = 0  # 列索引
        for cell in row:
            cell_index += 1  # 列索引自增, 从1开始
            if not cell:  # 跳过空白
                continue
            char = dec_to_alphanumeric(cell_index)  # 列索引转字母
            pos = f"{char}{row_index}"  # 拼接行列坐标
            cell_obj = sheet[pos]  # 根据行列坐标获取 cell 单元格对象
            # 对包含指定字符的单元格进行颜色填充
            if "迟到" in cell:
                cell_obj.fill = openpyxl.styles.PatternFill(start_color='FFEE0808', end_color='FFEE0808',
                                                            fill_type='solid')
                cell_obj.comment = openpyxl.comments.Comment("批注", "Comment Author")
            if "缺卡" in cell:
                cell_obj.fill = openpyxl.styles.PatternFill(start_color='FFFFFF00', end_color='FFFFFF00',
                                                            fill_type='solid')
                cell_obj.comment = openpyxl.comments.Comment("批注", "Comment Author")
            if "早退" in cell:
                cell_obj.fill = openpyxl.styles.PatternFill(start_color='FFFFFF00', end_color='FFFFFF00',
                                                            fill_type='solid')
                cell_obj.comment = openpyxl.comments.Comment("批注", "Comment Author")
            if "假" in cell:
                cell_obj.fill = openpyxl.styles.PatternFill(start_color='FFF4AF85', end_color='FFF4AF85',
                                                            fill_type='solid')
                cell_obj.comment = openpyxl.comments.Comment("批注", "Comment Author")
            if "出差" in cell:
                cell_obj.fill = openpyxl.styles.PatternFill(start_color='FFCCFFFF', end_color='FFCCFFFF',
                                                            fill_type='solid')
                cell_obj.comment = openpyxl.comments.Comment("批注", "Comment Author")

    # 保存并关闭
    workbook.save(excel_output_path)  # 保存文件
    workbook.close()  # 关闭文件


# 设置常量 =============================================================================================================
IS_OVER = False
VIDEO_DIR = ""


# 自动录屏 =============================================================================================================


# GUI 界面 =============================================================================================================
class GUI:
    def __init__(self):
        self.root = tk.Tk()
        # 设置窗口信息
        WIDTH = 400
        HEIGHT = 300
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width / 2) - (WIDTH / 2)
        y = (screen_height / 2) - (HEIGHT / 2)
        self.root.geometry(f"{WIDTH}x{HEIGHT}+{int(x)}+{int(y)}")  # 窗口尺寸, 坐标
        self.root.title("飞书自动考勤机器人")

    def set_grid(self):
        # 建立布局
        for i in range(2):  # 3列
            self.root.columnconfigure(i, weight=1)
        for i in range(6):  # 6行
            self.root.rowconfigure(i, weight=1)

        # 创建背景
        image = Image.open(r"F:\@#\DevOps - development& operations\Code - Project：pyhton& processing& c\pyJobFeishuCrawler\1.jpg")  # 确保图片路径正确
        image = ImageTk.PhotoImage(image)

        # 创建一个Label小部件来展示图片，并使用pack()方法将它放置在窗口的左上角
        label = tk.Label(self.root, image=image)
        label.image = image
        label.grid(row=3, column=3)

        # 创建标签
        label1 = tk.Label(self.root, text="开始日期:")
        label1.grid(row=1, column=0)
        label2 = tk.Label(self.root, text="结束日期:")
        label2.grid(row=2, column=0)
        label3 = tk.Label(self.root, text="文件输出目录:")
        label3.grid(row=3, column=0)

        # 创建输入框
        self.entry1 = tk.Entry(self.root, textvariable=tk.StringVar(value="20200202"))
        self.entry1.grid(row=1, column=1)
        self.entry2 = tk.Entry(self.root, textvariable=tk.StringVar(value="20200202"))
        self.entry2.grid(row=2, column=1)
        self.entry3 = tk.Entry(self.root, textvariable=tk.StringVar(value=r"D:\飞书考勤统计"))
        self.entry3.grid(row=3, column=1)

        # 创建按钮并绑定事件
        button1 = tk.Button(self.root, text="开始执行", command=self._func_start)
        button1.grid(row=4, column=0)
        button2 = tk.Button(self.root, text="退出程序", command=self.root.destroy)
        button2.grid(row=4, column=1)

    def display(self):
        self.set_grid()
        self.root.mainloop()
        # self.get_video()

    def _func_start(self):
        global IS_OVER, VIDEO_DIR, thread_2
        start_date = self.entry1.get()
        end_date = self.entry2.get()
        output_dir = self.entry3.get()

        if not os.path.isdir(output_dir):
            tk.messagebox.showinfo(title='提示', message=f'文件目录\n\n{output_dir}\n\n不存在\n\n请手动创建文件目录')
            return False
        else:
            VIDEO_DIR = output_dir
            thread_2.start()  # 线程1启动

        pattern = r'^\d{4}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$'  # 正则表达式，匹配一个字母
        if bool(re.match(pattern, start_date)) and bool(re.match(pattern, end_date)):
            excel_download_path = os.path.join(output_dir, f"{start_date}-{end_date}导出考勤.xlsx")
            excel_output_path = os.path.join(output_dir, f"{start_date}-{end_date}汇总统计.xlsx")
        else:
            tk.messagebox.showinfo(title='提示', message='日期格式不正确, 请重新输入\n\n日期格式: 20220101')
            return False

        try:
            download_excel(excel_download_path=excel_download_path, excel_output_path=excel_output_path,
                           start_date=start_date, end_date=end_date)
            change_excel(excel_download_path=excel_download_path, excel_output_path=excel_output_path,
                         start_date=start_date, end_date=end_date)
            tk.messagebox.showinfo(title='提示', message='执行成功')

        except Exception as e:
            tk.messagebox.showinfo(title='提示', message=f'执行失败\n{e}')

        IS_OVER = True


def get_video():
    # 录屏
    global VIDEO_DIR
    temp = tk.Tk()
    screen_size = (temp.winfo_screenwidth(), temp.winfo_screenheight())  # 获取屏幕分辨率
    fourcc = cv2.VideoWriter_fourcc(*"XVID")  # 设置视频编码器
    # fourcc = cv2.VideoWriter_fourcc(*"X264")  # 设置视频编码器
    form_time = time.strftime("%Y-%m-%d %H：%M：%S", time.localtime())  # 当前时间

    # 创建视频写入对象
    out = cv2.VideoWriter(os.path.join(VIDEO_DIR, f"{form_time}.mp4"), fourcc, 20.0, screen_size)  #
    print(os.path.join(VIDEO_DIR, f"{form_time}.avi"))
    # 开始录屏
    while not IS_OVER:
        # 获取屏幕截图
        img = pyautogui.screenshot()

        # 将截图转换为OpenCV格式
        frame = np.array(img)
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # 写入视频
        out.write(frame)
    # 释放资源
    out.release()
    cv2.destroyAllWindows()


def main():
    gui_obj = GUI()
    gui_obj.display()


thread_1 = Thread(target=main)  # 线程2：get_video
thread_2 = Thread(target=get_video)  # 线程1：mainloop

thread_1.start()  # 线程1启动

# exit()
